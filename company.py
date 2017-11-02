# -*- coding: utf-8 -*-
import PIL.Image as image
import PIL.ImageChops as imagechops
import time,re,io,urllib,random
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys  
from bs4 import BeautifulSoup
import requests
import os
from openpyxl import Workbook
from openpyxl import load_workbook


class spider():
    def __init__(self,filename=r'重庆公司信息(工商局版)'):
        self.path=os.path.join(os.path.abspath('.'),filename+'.xlsx')
        if os.path.exists(self.path):
            self.wb = load_workbook(self.path)
            self.ws1 = self.wb.get_sheet_by_name('公司信息')
            self.ws2 = self.wb.get_sheet_by_name('未爬到的公司名字')
            self.ws3 = self.wb.get_sheet_by_name('超过五页关键字')
            self.ws4 = self.wb.get_sheet_by_name('未验证成功的关键字')
        else:
            self.wb = Workbook()
            self.ws1=self.wb.create_sheet('公司信息',0)
            self.ws2=self.wb.create_sheet('未爬到的公司名字',1)
            self.ws3=self.wb.create_sheet('超过五页关键字',2)
            self.ws4=self.wb.create_sheet('未验证成功的关键字',3)
        self.ws1.cell(row =1, column =1, value='公司名字')
        self.ws1.cell(row =1, column =2, value='统一社会信誉代码')
        self.ws1.cell(row =1, column =3, value='公司类型')
        self.ws1.cell(row =1, column =4, value='法定代表人')
        self.ws1.cell(row =1, column =5, value='注册资本')
        self.ws1.cell(row =1, column =6, value='成立日期')
        self.ws1.cell(row =1, column =7, value='核准日期')
        self.ws1.cell(row =1, column =8, value='登记机关')
        self.ws1.cell(row =1, column =9, value='登记状态')
        self.ws1.cell(row =1, column =10, value='营业场所')
        self.ws1.cell(row =1, column =11, value='经营范围')
        
        self.ws2.cell(row =1, column =1, value='公司名字')
        
        self.ws3.cell(row =1, column =1, value='关键字')
        
        self.ws4.cell(row =1, column =1, value='关键字')
        self.wb.save(self.path)
    def get_merge_image(self,filename,location_list):
        im = image.open(filename)
        new_im = image.new('RGB', (260,116))
        im_list_upper=[]
        im_list_down=[]
        for location in location_list:
            if location['y']==-58:
                im_list_upper.append(im.crop((abs(location['x']),58,abs(location['x'])+10,166)))
            if location['y']==0:
                im_list_down.append(im.crop((abs(location['x']),0,abs(location['x'])+10,58)))
        new_im = image.new('RGB', (260,116))
        x_offset = 0
        for im in im_list_upper:
            new_im.paste(im, (x_offset,0))
            x_offset += im.size[0]
        x_offset = 0
        for im in im_list_down:
            new_im.paste(im, (x_offset,58))
            x_offset += im.size[0]    
        return new_im
    def get_image(self,driver,div):
        #'''    下载并还原图片    :driver:webdriver    :div:图片的div    '''
        #找到图片所在的div
        header= {
            'Connection':'close',
            'User-Agent':'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US; rv:1.9.1.6) Gecko/20091201 Firefox/3.5.6'
            }
        background_images=driver.find_elements_by_xpath(div)
        location_list=[]
        imageurl=''
        for background_image in background_images:
            location={}
            #在html里面解析出小图片的url地址，还有长高的数值
            location['x']=int(re.findall("background-image: url\(\"(.*)\"\); background-position: (.*)px (.*)px;",background_image.get_attribute('style'))[0][1])
            location['y']=int(re.findall("background-image: url\(\"(.*)\"\); background-position: (.*)px (.*)px;",background_image.get_attribute('style'))[0][2])
            imageurl=re.findall("background-image: url\(\"(.*)\"\); background-position: (.*)px (.*)px;",background_image.get_attribute('style'))[0][0]
            location_list.append(location)
            imageurl=imageurl.replace("webp","jpg")
            #print(imageurl)
            try:
				#req=urllib.request.urlopen(imageurl)
				#req.read()
                jpgfile=io.BytesIO(requests.get(imageurl,headers=header,timeout=(3.05, 27)).content)
            except Exception as e:
                #req.close()
                print('获取滑动验证图片失败,重新获取!')
                print('*'*80)
                print(e)
                print('*'*80)
                return False
            #重新合并图片
			#req.close()
            image=self.get_merge_image(jpgfile,location_list )
        return image
            
    def is_similar(self,image1,image2,x,y):
        #'''    对比RGB值    '''
        pixel1=image1.getpixel((x,y))
        pixel2=image2.getpixel((x,y))
        for i in range(0,3):
            if abs(pixel1[i]-pixel2[i])>=50:
                return False
        return True
    def get_diff_location(self,image1,image2):
        #'''    计算缺口的位置    '''
        i=0
        for i in range(0,260):
            for j in range(0,116):
                if self.is_similar(image1,image2,i,j)==False:
                    return  i,j
    def get_track(self,length):
        list=[]
        #间隔通过随机范围函数来获得,每次移动一步或者两步
        x=random.randint(2,4)
        #生成轨迹并保存到list内
        total=length-5
        #step1=total//2
        step1=total//4*3
        list.append(step1)
        step2=(total-step1)//4*3
        list.append(step2)
        total2=total-step1-step2
        while total2-x>5:
            list.append(x)
            total2=total2-x
            x=random.randint(2,5)
        list.append(total2)
        return list
    def hk_veriy(self,driver):
        WebDriverWait(driver, 30).until(lambda the_driver: the_driver.find_element_by_xpath("//div[@class='gt_cut_bg gt_show']").is_displayed())
        WebDriverWait(driver, 30).until(lambda the_driver: the_driver.find_element_by_xpath("//div[@class='gt_cut_fullbg gt_show']").is_displayed())
        WebDriverWait(driver, 30).until(lambda the_driver: the_driver.find_element_by_xpath("//div[@class='gt_slider_knob gt_show']").is_displayed())
        image1_xpath="//div[@class='gt_cut_bg gt_show']/div"
        image2_xpath="//div[@class='gt_cut_fullbg gt_show']/div"
        image1=self.get_image(driver,image1_xpath)
        image2=self.get_image(driver,image2_xpath)
        if not(image1 and image2):
            return False
        #圆球位置
        element=driver.find_element_by_xpath("//div[@class='gt_slider_knob gt_show']")
        location=element.location
        #缺口位置
        diff_location=self.get_diff_location(image1,image2)
        #x轨迹
        track_list=self.get_track(diff_location[0])
        #按住圆球
        ActionChains(driver).click_and_hold(on_element=element).perform()
        time.sleep(0.15)
        #x，y偏移
        t1=time.time()
        #注意鼠标抖动范围5像素
        i=0
        j=1
        f=0
        for track in track_list:
            yoffset=random.randint(22,25)
            if len(track_list)<7:
                j=1.5
            if len(track_list)>15:
                j=0.5
            if i>len(track_list)//len(track_list)*(len(track_list)-1):
                k=4
            elif i>len(track_list)//2:
                k=2
            else:
                k=1
            i=i+1
            ActionChains(driver).move_to_element_with_offset(to_element=element, xoffset=track+22, yoffset=yoffset).perform()
            time.sleep(random.randint(40,50)/500*k*j)
        ActionChains(driver).move_to_element_with_offset(to_element=element, xoffset=22+2, yoffset=yoffset).perform()
        time.sleep(random.randint(50,60)/500)
        ActionChains(driver).move_to_element_with_offset(to_element=element, xoffset=22-2, yoffset=yoffset).perform()
        time.sleep(random.randint(50,60)/500)
        ActionChains(driver).release(on_element=element).perform()
        try:
            WebDriverWait(driver, 15).until(lambda the_driver: the_driver.find_element_by_xpath("//div[@class='gt_ajax_tip gt_success']").is_displayed())
            #driver.find_element_by_xpath("//div[@class='gt_ajax_tip gt_success']").is_displayed()
        except Exception as e:
            try:
                WebDriverWait(driver, 10).until(lambda the_driver: the_driver.find_element_by_xpath("//div[@class='gt_cut_bg gt_show']").is_displayed())
            except Exception as e1:
                return True
            #print(e)
            print('滑动验证失败,重新验证。')
            return False
        return True
    def get_all_company(self,driver,cookies,key_word):
        html=driver.page_source
        soup=BeautifulSoup(html,'html.parser')
        page_form=soup.find_all('div',{'class':'pagination'})
        print(page_form)
        if page_form:
            page_a_list=page_form[0].find_all('a')
            a_list=soup.find_all('a',{'class':'search_list_item db'})
            #print(a_list)
            for link in a_list:
                name=link.find('h1').get_text()
                self.html_parser(link.get('href'),cookies,name)
            if len(page_a_list)>2:
                if len(page_a_list)==6:
                    self.ws3.cell(row =self.ws3.max_row+1, column =1, value=key_word[-1])
                    self.wb.save(self.path)
                for page in range(len(page_a_list)-2):
                    #下一页
                    old_url=driver.current_url
                    if page:
                        k=2
                    else:
                        k=0
                    try:
                        driver.find_element_by_xpath('/html/body/div[5]/div[3]/div[2]/form/a[{next_page}]'.format(next_page=len(page_a_list)-1+k)).click()
                    except:
                        try:
                            driver.find_element_by_link_text('{page}'.format(page=page+2))
                        except:
                            break
                    i=0
					#翻页等待30秒,每10秒重新翻页!
                    while driver.current_url==old_url:
                        i=i+1
                        time.sleep(1)
                        if i%10:
                            try:
                                driver.find_element_by_link_text('{page}'.format(page=page+2))
                            except:
                                pass
                        if i==30:
                            print('翻页失败,跳出循环爬取下一页!')
                            break
                    if 'corp-query-search' not in driver.current_url:
                        driver.quit()
                        return
                    html=driver.page_source
                    soup=BeautifulSoup(html,'html.parser')
                    a_list=soup.find_all('a',{'class':'search_list_item db'})
                    for link in a_list:
                        name=link.find('h1').get_text()
                        self.html_parser(link.get('href'),cookies,name)
        else:
            print('该关键字没有匹配的公司!')
        print('关闭浏览器')
        driver.quit()
        print('完成关闭浏览器')
    def html_parser(self,url,cookies,name):
        url='http://www.gsxt.gov.cn'+url
        header= {
            'Connection':'close',
            'User-Agent':'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US; rv:1.9.1.6) Gecko/20091201 Firefox/3.5.6'
            }
        result={}
        try:
            print('开始获取公司:{name}————————>'.format(name=name))
            res=requests.get(url,headers=header,cookies=cookies,timeout=(3.05, 27))
            print('开始解析公司:{name}————————>'.format(name=name))
            soup = BeautifulSoup(res.content,'html.parser')
            fullName_tag=soup.find(attrs={'class':'fullName'})
            if fullName_tag:
                info_div=soup.find(attrs={'class':'overview'})
                info_dl=info_div.find_all('dl')
                result['公司名字: ']=fullName_tag.get_text().replace('\t','').replace('\r','').replace('\n','').strip()
                for dl in info_dl:
                    print(dl)
                    dict_name=dl.find_all('dt')[0].get_text().replace('\t','').replace('\r','').replace('\n','').strip()
                    dict_value=dl.find_all('dd')[0].get_text().replace('\t','').replace('\r','').replace('\n','').strip()
                    result[dict_name]=dict_value
                    #print(result)
                self.write_info_excel(result)
            else:
                #print('not dound',name)
                print('解析公司失败:{name}————————>'.format(name=name))
                self.ws2.cell(row =self.ws2.max_row+1, column =1, value=name)
                self.wb.save(self.path)
        except Exception as e:
            print('解析公司失败:{name}————————>'.format(name=name))
            print('*'*80)
            print(e)
            print('*'*80)
            self.ws2.cell(row =self.ws2.max_row+1, column =1, value=name)
            self.wb.save(self.path)
    def write_info_excel(self,result):
        row_max=self.ws1.max_row
        for name,value in result.items():
            if name=='公司名字: ':
                self.ws1.cell(row =row_max+1, column =1, value=value)
            if name=='统一社会信用代码：' or name=='注册号：':
                self.ws1.cell(row =row_max+1, column =2, value=value)
            if name=='类型：':
                self.ws1.cell(row =row_max+1, column =3, value=value)
            if name=='法定代表人：' or name=='负责人：':
                self.ws1.cell(row =row_max+1, column =4, value=value)
            if name=='注册资本：':
                self.ws1.cell(row =row_max+1, column =5, value=value)
            if name=='成立日期：':
                self.ws1.cell(row =row_max+1, column =6, value=value)
            if name=='核准日期：':
                self.ws1.cell(row =row_max+1, column =7, value=value)
            if name=='登记机关：':
                self.ws1.cell(row =row_max+1, column =8, value=value)
            if name=='登记状态：':
                self.ws1.cell(row =row_max+1, column =9, value=value)
            if name=='营业场所：' or name=='住所：':
                self.ws1.cell(row =row_max+1, column =10, value=value)
            if name=='经营范围：':
                self.ws1.cell(row =row_max+1, column =11, value=value)
        self.wb.save(self.path)     
    def search_keyword_parser_write(self,keyword):
        driver=webdriver.Firefox()
        time.sleep(2)
        try:
            driver.get("http://www.gsxt.gov.cn/index.html")
            try:
                WebDriverWait(driver, 30).until(lambda the_driver: the_driver.find_element_by_id("keyword").is_displayed())
                WebDriverWait(driver, 30).until(lambda the_driver: the_driver.find_element_by_id("btn_query").is_displayed())
            except:
                driver.get("http://www.gsxt.gov.cn/index.html")
                WebDriverWait(driver, 30).until(lambda the_driver: the_driver.find_element_by_id("keyword").is_displayed())
                WebDriverWait(driver, 30).until(lambda the_driver: the_driver.find_element_by_id("btn_query").is_displayed())
            driver.find_element_by_id('keyword').send_keys(keyword)
            driver.find_element_by_id('btn_query').click()
            old_url=driver.current_url
            t=0
            try:
                hk=self.hk_veriy(driver)
            except Exception as e:
                print(e)
                hk=False
            while not hk:
                time.sleep(2)
                t=t+1
                print('滑动验证失败,第{tt}次,重新认证!'.format(tt=t))
                if t==7:
                    print('滑动验证失败,超过7次,跳出关键字:{key}!'.format(key=keyword))
                    row_max=self.ws4.max_row
                    self.ws4.cell(row =row_max+1, column =11, value=keyword[2:])
                    self.wb.save(self.path)
                    driver.quit()
                    return
                try:
                    hk=self.hk_veriy(driver)
                except Exception as e:
                    print(e)
                    hk=False
            t=0
            while driver.current_url==old_url:
                print('等待跳转成功!')
                time.sleep(1)
                t=t+1
                if t==50:
                    print('跳转失败,爬取下一个关键字!')
                    driver.quit()
                    return
            if 'corp-query-search' not in driver.current_url:
                print('跳转失败,爬取下一个关键字!')
                driver.quit()
                return
        except Exception as e:
            driver.quit()
            print(keyword+'---搜索失败!')
            print('*'*80)
            print(e)
            print('*'*80)
            return
        cookies={}
        for item in driver.get_cookies():
            cookies[item["name"]]=item["value"]
        self.get_all_company(driver,cookies,keyword)
if __name__=='__main__':
    ss=spider(r'重庆公司信息(工商局版)')
    #for x in range(19968,40869):
    list=['的','一','是','在','不','了','有','和','人','这','中','大','为','上','个','国','我','以','要','他','时','来','用','们','生','到','作','地','于','出','就','分','对','成','会','可','主','发','年','动','同','工','也','能','下','过','子','说','产','种','面','而','方','后','多','定','行','学','法','所','民','得','经','十','三','之','进','着','等','部','度','家','电','力','里','如','水','化','高','自','二','理','起','小','物','现','实','加','量','都','两','体','制','机','当','使','点','从','业','本','去','把','性','好','应','开','它','合','还','因','由','其','些','然','前','外','天','政','四','日','那','社','义','事','平','形','相','全','表','间','样','与','关','各','重','新','线','内','数','正','心','反','你','明','看','原','又','么','利','比','或','但','质','气','第','向','道','命','此','变','条','只','没','结','解','问','意','建','月','公','无','系','军','很','情','者','最','立','代','想','已','通','并','提','直','题','党','程','展','五','果','料','象','员','革','位','入','常','文','总','次','品','式','活','设','及','管','特','件','长','求','老','头','基','资','边','流','路','级','少','图','山','统','接','知','较','长','将','组','见','计','别','她','手','角','期','根','论','运','农','指','几','九','区','强','放','决','西','被','干','做','必','战','先','回','则','任','取','据','处','队','南','给','色','光','门','即','保','治','北','造','百','规','热','领','七','海','地','口','东','导','器','压','志','世','金','增','争','济','阶','油','思','术','极','交','受','联','什','认','六  共','权','收','证','改','清','已','美','再','采','转','更','单','风','切','打','白','教','速','花','带','安','场','身','车','例','真','务','具','万','每','目','至','达','走','积','示','议','声','报','斗','完','类','八','离','华','名','确','才','科','张','信','马','节','话','米','整','空','元','况','今','集','温','传','土','许','步','群','广','石','记','需','段','研','界','拉','林','律','叫','且','究','观','越','织','装','影','算','低','持','音','众','书','布','复','容','儿','须','际','商','非','验','连','断','深','难','近','矿','千','周','委','素','技','备','半','办','青','省','列','习','响','约','支','般','史','感','劳','便','团','往','酸','历','市','克','何','除','消','构','府','称','太','准','精','值','号','率','族','维','划','选','标','写','存','候','毛','亲','快','效','斯','院','查','江','型','眼','王','按','格','养','易','置','派','层','片','始','却','专','状','育','厂','京','识','适','属','圆','包','火','住','调','满','县','局','照','参','红','细','引','听','该','铁','价','严','首','底','液','官','德','调','随','病','苏','失','尔','死','讲','配','女','黄','推','显','谈','罪','神','艺','呢','席','含','企','望','密','批','营','项','防','举','球','英','氧','势','告','李','台','落','木','帮','轮','破','亚','师','围','注','远','字','材','排','供','河','态','封','另','施','减','树','溶','怎','止','案','言','士','均','武','固','叶','鱼','波','视','仅','费','紧','爱','左','章','早','朝','害','续','轻','服','试','食','充','兵','源','判','护','司','足','某','练','差','致','板','田','降','黑','犯','负','击','范','继','兴','似','余','坚','曲','输','修','的','故','城','夫','够','送','笔','船','占','右','财','吃','富','春','职','觉','汉','画','功','巴','跟','虽','杂','飞','检','吸','助','升','阳','互','初','创','抗','考','投','坏','策','古','径','换','未','跑','留','钢','曾','端','责','站','简','述','钱','副','尽','帝','射','草','冲','承','独','令','限','阿','宣','环','双','请','超','微','让','控','州','良','轴','找','否','纪','益','依','优','顶','础','载','倒','房','突','坐','粉','敌','略','客','袁','冷','胜','绝','析','块','剂','测','丝','协','重','诉','念','陈','仍','罗','盐','友','洋','错','苦','夜','刑','移','频','逐','靠','混','母','短','皮','终','聚','汽','村','云','哪','既','距','卫','停','烈','央','察','烧','行','迅','境','若','印','洲','刻','括','激','孔','搞','甚','室','待','核','校','散','侵','吧','甲','游','久','菜','味','旧','模','湖','货','损','预','阻','毫','普','稳','乙','妈','植','息','扩','银','语','挥','酒','守','拿','序','纸','医','缺','雨','吗','针','刘','啊','急','唱','误','训','愿','审','附','获','茶','鲜','粮','斤','孩','脱','硫','肥','善','龙','演','父','渐','血','欢','械','掌','歌','沙','著','刚','攻','谓','盾','讨','晚','粒','乱','燃','矛','乎','杀','药','宁','鲁','贵','钟','煤','读','班','伯','香','介','迫','句','丰','培','握','兰','担','弦','蛋','沉','假','穿','执','答','乐','谁','顺','烟','缩','征','脸','喜','松','脚','困','异','免','背','星','福','买','染','井','概','慢','怕','磁','倍','祖','皇','促','静','补','评','翻','肉','践','尼','衣','宽','扬','棉','希','伤','操','垂','秋','宜','氢','套','笔','督','振','架','亮','末','宪','庆','编','牛','触','映','雷','销','诗','座','居','抓','裂','胞','呼','娘','景','威','绿','晶','厚','盟','衡','鸡','孙','延','危','胶','还','屋','乡','临','陆','顾','掉','呀','灯','岁','措','束','耐','剧','玉','赵','跳','哥','季','课','凯','胡','额','款','绍','卷','齐','伟','蒸','殖','永','宗','苗','川','炉','岩','弱','零','杨','奏','沿','露','杆','探','滑','镇','饭','浓','航','怀','赶','库','夺','伊','灵','税','了','途','灭','赛','归','召','鼓','播','盘','裁','险','康','唯','录','菌','纯','借','糖','盖','横','符','私','努','堂','域','枪','润','幅','哈','竟','熟','虫','泽','脑','壤','碳','欧','遍','侧','寨','敢','彻','虑','斜','薄','庭','都','纳','弹','饲','伸','折','麦','湿','暗','荷','瓦','塞','床','筑','恶','户','访','塔','奇','透','梁','刀','旋','迹','卡','氯','遇','份','毒','泥','退','洗','摆','灰','彩','卖','耗','夏','择','忙','铜','献','硬','予','繁','圈','雪','函','亦','抽','篇','阵','阴','丁','尺','追','堆','雄','迎','泛','爸','楼','避','谋','吨','野','猪','旗','累','偏','典','馆','索','秦','脂','潮','爷','豆','忽','托','惊','塑','遗','愈','朱','替','纤','粗','倾','尚','痛','楚','谢','奋','购','磨','君','池','旁','碎','骨','监','捕','弟','暴','割','贯','殊','释','词','亡','壁','顿','宝','午','尘','闻','揭','炮','残','冬','桥','妇','警','综','招','吴','付','浮','遭','徐','您','摇','谷','赞','箱','隔','订','男','吹','乐','园','纷','唐','败','宋','玻','巨','耕','坦','荣','闭','湾','键','凡','驻','锅','救','恩','剥','凝','碱','齿','截','炼','麻','纺','禁','废','盛','版','缓','净','睛','昌','婚','涉','筒','嘴','插','岸','朗','庄','街','藏','姑','贸','腐','奴','啦','惯','乘','伙','恢','匀','纱','扎','辩','耳','彪','臣','亿','璃','抵','脉','秀','萨','俄','网','舞','店','喷','纵','寸','汗','挂','洪','着','贺','闪','柬','爆','烯','津','稻','墙','软','勇','像','滚','厘','蒙','芳','肯','坡','柱','荡','腿','仪','旅','尾','轧','冰','贡','登','黎','削','钻','勒','逃','障','氨','郭','峰','币','港','伏','轨','亩','毕','擦','莫','刺','浪','秘','援','株','健','售','股','岛','甘','泡','睡','童','铸','汤','阀','休','汇','舍','牧','绕','炸','哲','磷','绩','朋','淡','尖','启','陷','柴','呈','徒','颜','泪','稍','忘','泵','蓝','拖','洞','授','镜','辛','壮','锋','贫','虚','弯','摩','泰','幼','廷','尊','窗','纲','弄','隶','疑','氏','宫','姐','震','瑞','怪','尤','琴','循','描','膜','违','夹','腰','缘','珠','穷','森','枝','竹','沟','催','绳','忆','邦','剩','幸','浆','栏','拥','牙','贮','礼','滤','钠','纹','弹','罢','拍','咱','喊','袖','埃','勤','罚','焦','潜','伍','墨','欲','缝','姓','刊','饱','仿','奖','铝','鬼','丽','跨','默','挖','链','扫','喝','袋','炭','污','幕','诸','弧','励','梅','奶','洁','灾','舟','鉴','苯','讼','抱','毁','率','懂','寒','智','埔','寄','届','跃','渡','挑','丹','艰','贝','碰','拔','爹','戴','码','梦','芽','熔','赤','渔','哭','敬','颗','奔','藏','铅','熟','仲','虎','稀','妹','乏','珍','申','桌','遵','允','隆','螺','仓','魏','锐','晓','氮','兼','隐','碍','赫','拨','忠','肃','缸','牵','抢','博','巧','壳','兄','杜','讯','诚','碧','祥','柯','页','巡','矩','悲','灌','龄','伦','票','寻','桂','铺','圣','恐','恰','郑','趣','抬','荒','腾','贴','柔','滴','猛','阔','辆','妻','填','撤','储','签','闹','扰','紫','砂','递','戏','吊','陶','伐','喂','疗','瓶','婆','抚','臂','摸','忍','虾','蜡','邻','胸','巩','挤','偶','弃','槽','劲','乳','邓','吉','仁','烂','砖','租','乌','舰','伴','瓜','浅','丙','暂','燥','橡','柳','迷','暖','牌','纤','秧','胆','详','簧','踏','瓷','谱','呆','宾','糊','洛','辉','愤','竞','隙','怒','粘','乃','绪','肩','籍','敏','涂','熙','皆','侦','悬','掘','享','纠','醒','狂','锁','淀','恨','牲','霸','爬','赏','逆','玩','陵','祝','秒','浙','貌','役','彼','悉','鸭','着','趋','凤','晨','畜','辈','秩','卵','署','梯','炎','滩','棋','驱','筛','峡','冒','啥','寿','译','浸','泉','帽','迟','硅','疆','贷','漏','稿','冠','嫩','胁','芯','牢','叛','蚀','奥','鸣','岭','羊','凭','串','塘','绘','酵','融','盆','锡','庙','筹','冻','辅','摄','袭','筋','拒','僚','旱','钾','鸟','漆','沈','眉','疏','添','棒','穗','硝','韩','逼','扭','侨','凉','挺','碗','栽','炒','杯','患','馏','劝','豪','辽','勃','鸿','旦','吏','拜','狗','埋','辊','掩','饮','搬','骂','辞','勾','扣','估','蒋','绒','雾','丈','朵','姆','拟','宇','辑','陕','雕','偿','蓄','崇','剪','倡','厅','咬','驶','薯','刷','斥','番','赋','奉','佛','浇','漫','曼','扇','钙','桃','扶','仔','返','俗','亏','腔','鞋','棱','覆','框','悄','叔','撞','骗','勘','旺','沸','孤','粘','吐','孟','渠','屈','疾','妙','惜','仰','狠','胀','谐','抛','霉','桑','岗','嘛','衰','盗','渗','脏','赖','涌','甜','曹','阅','肌','哩','厉','烃','纬','毅','昨','伪','症','煮','叹','钉','搭','茎','笼','酷','偷','弓','锥','恒','杰','坑','鼻','翼','纶','叙','狱','逮','罐','络','棚','抑','膨','蔬','寺','骤','穆','冶','枯','册','尸','凸','绅','坯','牺','焰','轰','欣','晋','瘦','御','锭','锦','丧','旬','锻','垄','搜','佛','扑','邀','亭','酯','迈','舒','脆','酶','闲','忧','酚','顽','羽','涨','卸','仗','陪','薄','辟','惩','杭','姚','肚','捉','飘','漂','昆','欺','吾','郎','烷','汁','呵','饰','萧','雅','邮','迁','燕','撒','姻','赴','宴','烦','削','债','帐','斑','铃','旨','醇','董','饼','雏','姿','拌','傅','腹','妥','揉','贤','拆','歪','葡','胺','丢','浩','徽','昂','垫','挡','览','贪','慰','缴','汪','慌','冯','诺','姜','谊','凶','劣','诬','耀','昏','躺','盈','骑','乔','溪','丛','卢','抹','易','闷','咨','刮','驾','缆','悟','摘','铒','掷','颇','幻','柄','惠','惨','佳','仇','腊','窝','涤','剑','瞧','堡','泼','葱','罩','霍','捞','胎','苍','滨','俩','捅','湘','砍','霞','邵','萄','疯','淮','遂','熊','粪','烘','宿','档','戈','驳','嫂','裕','徙','箭','捐','肠','撑','晒','辨','殿','莲','摊','搅','酱','屏','疫','哀','蔡','堵','沫','皱','畅','叠','阁','莱','敲','辖','钩','痕','坝','巷','饿','祸','丘','玄','溜','曰','逻','彭','尝','卿','妨','艇','吞','韦','怨','矮','歇','郊','禄','捻','漠','粹','颠','宏','冤','肪','饥','呵','仙','押','挨','醛','娃','拾','没','佩','勿','吓','讹','侯','恋','夕','锌','篡','戚','淋','蓬','岂','釉','兆','泊','魂','拘','亡','杠','摧','氟','颂','浑','凌','铀','诱','犁','谴','颁','舶','扯','嘉','萌','犹','滋','焊','舌','匹','媳','肺','掠','酿','烹','疲','驰','鸦','窄','辱','狭','朴','遣','菲','奸','韧','辣','拳','秆','卧','醉','竭','茅','墓','矣','哎','艳','敦','舆','缔','雇','尿','葬','履','契','禽','渣','衬','躲','赔','咸','溉','贼','醋','堤','抖','妃','裤','廉','晴','挽','掀','茫','丑','亥','拦','悠','阐','慧','佐','奇','竖','孝','柜','麟','绣','遥','逝','愁','肖','昭','芬','逢','窑','捷','圜','盲','闸','宙','辐','披','账','狼','幽','绸','蜂','慎','餐','酬','誓','惟','叉','弥','址','帜','芝','砌','唉','仆','涛','臭','翠','盒','劫','慨','炳','阖','寂','椒','倘','拓','畏','喉','巾','颈','垦','拚','兽','蔽','芦','乾','爽','窃','谭','挣','崩','模','褐','传','翅','儒','伞','晃','谬','胚','剖','凑','眠','浊','霜','礁','蔑','抄','闯','洒','碑','蓉','耶','猜','蹲','壶','唤','澳','锯','郡','玲','绵','纽','梳','掏','吁','锤','鼠','穴','椅','殷','遮','吵','萍','厌','畜','俱','夸','吕','囊','捧','雌','闽','饶','瞬','郁','哨','凿','朝','俺','浒','茂','肝','勋','盯','籽','耻','菊','滥','稼','戒','奈','帅','鞭','蚕','镁','询','跌','烤','坛','宅','笛','鄂','蛮','颤','棍','睁','鼎','岌','降','侍','藩','嚷','匪','岳','糟','缠','迪','泄','卑','氛','堪','萝','盛','碘','缚','悦','澄','甫','攀','屠','溢','拱','晰','携','朽','吟','菱','谦','凹','俊','芒','盼','婶','艘','酰','趁','唇','挫','羞','浴','疼','萎','肴','愚','肿','刨','绞','枢','嫁','慕','舱','铲','苹','豫','谕','迭','潘','顷','翁','榜','匠','欠','茬','畴','胃','沾','踪','弊','哼','鹏','歧','桐','沃','悼','惑','溃','蔗','荐','潭','孢','露','诊','庸','聪','嫌','厨','庞','祁','钳','肆','梭','赠','崖','篮','颖','甸','藻','捣','且','撕','诏','贞','赐','慈','炕','胖','兹','差','琼','锈','汛','卓','棵','馈','挠','灶','婴','蒂','肤','衫','沥','仑','勉','沪','逸','蜜','浦','嗓','晕','膏','祭','赢','艾','扮','鹅','怜','蒲','兔','孕','呖','蘖','挪','淑','谣','惧','廊','缅','俘','骄','膀','陡','宰','诞','峻','恼','腺','猎','涡','夷','愉','魔','铵','葛','贾','似','荫','哟','脊','钞','苛','锰','椭','镶','杏','溴','倚','滞','会','氓','捏','斩','傲','匆','僵','卤','烫','衍','榨','拢','裸','屑','咽','坊','舅','渴','翔','邪','拄','窖','猫','砌','钦','媒','脾','勺','柏','栅','噪','昼','耿','扁','辰','秤','得','贩','糕','梁','昙','衷','宦','扔','哇','诈','嘱','藤','卜','冈','悔','廓','皂','拐','氰','杉','玛','矢','寓','瓣','罕','垮','笋','淘','衔','称','恭','喇','帕','桉','秉','帘','铭','蛇','摔','斋','叭','帆','裸','俭','瘤','篷','砸','肢','辟','脖','瞪','暑','卜','竿','歼','笙','酮','蕴','哗','瞎','喀','刃','楔','喘','枚','嵌','挝','厢','粤','甩','拴','膝','恳','腕','娓','熄','锚','忌','愧','哦','荆','圃','骚','丸','蒜','毯','弗','俯','鹿','梢','屯','衙','轿','贱','垒','谅','踢','哑','滔','渥','饷','泳','棕','熬','搁','腈','梨','吻','樱','奠','捆','姨','柏','聘','惕','郓','绑','冀','裹','酥','寡','彦','稠','啡','钝','汝','擅','汰','鳙','埔','敞','嘿','逊','栋','谨','咖','鲤','雀','佣','庵','葫','贿','鳞','拼','搏','谎','塌','忉','腻','戊','怖','坟','禾','刹','嘻','桔','坎','拇','煽','狮','痒','曾','梗','寇','鹰','烛','哄','莽','雯','胳','龟','亟','糠','泌','坪','傻','什','喻','渊','蚌','跪','巷','涅','钊','譬','蕊','膛','侮','奕','枕','辫','况','扼','郝','寥','凄','厦','腥','钧','耦','蹄','戥','屁','诵','匈','桩','钓','涵','倦','袍','抒','屿','蹈','忿','敷','虹','聊','嗣','尉','灿','糙','蹬','嗯','姬','狡','笨','辜','僧','茨','讽','翰','枉','岐','枣','崭','焚','咕','猴','揽','涝','耍','趟','汹','咋','傍','镀','给','爵','虏','劈','璋','踩','瞅','迄','昔','汞','呱','诡','魄','祺','嘲','惶','赃','癌','咐','歉','扳','鄙','庐','聂','便','芡','躯','贬','煌','拧','隋','襄','淤','宠','炊','滇','謇','懒','栓','佑','憾','骆','裙','猖','兜','孵','痼','盥','曝','泣','絮','韵','眷','旷','噢','参','栖','盏','鳌','溅','煎','校','榴','暮','琪','淆','陛','巢','哒','吼','槐','唧','其','沛','乞','蜀','蜇','赚','捍','铰','幂','尧','咒','耽','叮','褂','焕','煞','雹','搓','釜','铬','拣','募','淹','瑰','鲢','茄','灼','邹','躬','觉','娇','焉','彰','鹤','琳','沦','畔','惹','庶','毙','皖','邢','禹','渍','绷','窜','翘','淫','箪','陌','膊','鞑','咳','玫','巫','拂','蕉','澜','赎','绥','锄','囱','赌','颊','缕','寅','躁','稚','庚','苟','氦','魁','珊','蜕','蛭','酌','逗','闺','蔓','撇','豌','朕','缉','襟','镍','桅','荧','侄','卒','佃','瞿','娶','饪','耸','乍','靶','痴','靖','扛','筐','韶','嚣','崔','蓿','岔','氘','娥','剿','霖','喃','搪','雍','裳','撰','豹','骏','慷']
    for x in range(0,200):
        ss.search_keyword_parser_write('重庆'+list[x])